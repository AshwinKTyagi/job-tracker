"""Applications and events as tabular output: `pandas.DataFrame` and CSV/XLSX files.

`build_dataframe` and `build_events_dataframe` produce the frozen DataFrame shapes
(`constants.EXPORT_COLUMNS` / `constants.EVENT_COLUMNS`, invariant I10) that M5 consumes.
`write_csv` and `write_xlsx` are the only functions in this module that touch the
filesystem, and only at the caller-supplied `path`.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Any

import pandas as pd

# openpyxl has no bundled/PyPI type stubs in this venv, and adding one means editing the
# mypy overrides in pyproject.toml, which is M0's file — so these are ignored with reason.
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from jobtrack.constants import EVENT_COLUMNS, EXPORT_COLUMNS
from jobtrack.errors import ExportError
from jobtrack.models import ApplicationRow, EventRow

_MAX_COLUMN_WIDTH = 60
"""Cap on autosized column width (characters) so one long body/subject can't blow out a sheet."""

_NULLABLE_STRING_DTYPE = "string"
"""pandas' nullable string extension dtype. Explicit everywhere so behavior does not depend
on the ambient pandas-3 default-string-dtype setting (CLAUDE.md's pandas-3 note)."""

_UTC_DATETIME_DTYPE = "datetime64[ns, UTC]"


def build_dataframe(applications: Sequence[ApplicationRow]) -> pd.DataFrame:
    """Applications to a DataFrame with exactly `EXPORT_COLUMNS`, in order (I10).

    Args:
        applications: Rows to tabulate, in the order they should appear.

    Returns:
        A DataFrame with tz-aware UTC datetime columns and explicit dtypes on every
        column. Empty input yields an empty frame with the same columns and dtypes as a
        populated one — callers never need to special-case emptiness.
    """
    columns: dict[str, pd.Series[Any]] = {
        "application_id": pd.Series(
            [a.application_id for a in applications], dtype=_NULLABLE_STRING_DTYPE
        ),
        "company": pd.Series([a.company for a in applications], dtype=_NULLABLE_STRING_DTYPE),
        "role": pd.Series([a.role for a in applications], dtype=_NULLABLE_STRING_DTYPE),
        "location": pd.Series([a.location for a in applications], dtype=_NULLABLE_STRING_DTYPE),
        "ats": pd.Series([a.ats for a in applications], dtype=_NULLABLE_STRING_DTYPE),
        "status": pd.Series([str(a.status) for a in applications], dtype=_NULLABLE_STRING_DTYPE),
        "applied_at": pd.Series([a.applied_at for a in applications], dtype=_UTC_DATETIME_DTYPE),
        "last_event_at": pd.Series(
            [a.last_event_at for a in applications], dtype=_UTC_DATETIME_DTYPE
        ),
        "last_event_type": pd.Series(
            [str(a.last_event_type) for a in applications], dtype=_NULLABLE_STRING_DTYPE
        ),
        "event_count": pd.Series([a.event_count for a in applications], dtype="int64"),
        "days_to_first_response": pd.Series(
            [a.days_to_first_response for a in applications], dtype="Int64"
        ),
        "days_since_last_event": pd.Series(
            [a.days_since_last_event for a in applications], dtype="int64"
        ),
        "needs_review": pd.Series([a.needs_review for a in applications], dtype="bool"),
    }
    df = pd.DataFrame(columns)
    return df.loc[:, list(EXPORT_COLUMNS)]


def build_events_dataframe(events: Sequence[EventRow]) -> pd.DataFrame:
    """Events to a long-format DataFrame with exactly `EVENT_COLUMNS`, in order.

    Args:
        events: Rows to tabulate, in the order they should appear.

    Returns:
        A DataFrame with tz-aware UTC `occurred_at` and explicit dtypes on every column.
        Empty input yields an empty frame with the same columns and dtypes as a
        populated one.
    """
    columns: dict[str, pd.Series[Any]] = {
        "application_id": pd.Series(
            [e.application_id for e in events], dtype=_NULLABLE_STRING_DTYPE
        ),
        "message_id": pd.Series([e.message_id for e in events], dtype=_NULLABLE_STRING_DTYPE),
        "event_type": pd.Series([str(e.event_type) for e in events], dtype=_NULLABLE_STRING_DTYPE),
        "occurred_at": pd.Series([e.occurred_at for e in events], dtype=_UTC_DATETIME_DTYPE),
        "confidence": pd.Series([e.confidence for e in events], dtype="float64"),
        "needs_review": pd.Series([e.needs_review for e in events], dtype="bool"),
        "subject": pd.Series([e.subject for e in events], dtype=_NULLABLE_STRING_DTYPE),
    }
    df = pd.DataFrame(columns)
    return df.loc[:, list(EVENT_COLUMNS)]


def write_csv(df: pd.DataFrame, path: Path) -> Path:
    """Write a DataFrame to UTF-8 CSV with ISO-8601 dates and no index.

    Args:
        df: The frame to write, typically from `build_dataframe`.
        path: Destination file path.

    Returns:
        The resolved path that was written.

    Raises:
        ExportError: `path` is not writable (missing parent directory, a directory given
            in place of a file, permission denied, etc.).
    """
    resolved = path.resolve()
    try:
        df.to_csv(resolved, index=False, date_format="%Y-%m-%dT%H:%M:%S%z", encoding="utf-8")
    except OSError as exc:
        raise ExportError(f"could not write CSV to {resolved}: {exc}") from exc
    return resolved


def write_xlsx(df: pd.DataFrame, path: Path, *, events: pd.DataFrame | None = None) -> Path:
    """Write a DataFrame (and optionally an events frame) to an .xlsx workbook.

    Sheet 'Applications' always holds `df`; sheet 'Events' is added only when `events` is
    given. Each sheet freezes the header row, autosizes its columns, and gets an autofilter.

    Args:
        df: The applications frame, typically from `build_dataframe`.
        path: Destination file path.
        events: Optional events frame, typically from `build_events_dataframe`.

    Returns:
        The resolved path that was written.

    Raises:
        ExportError: `path` is not writable, or openpyxl failed to build/save the workbook.
    """
    resolved = path.resolve()
    try:
        workbook = Workbook()
        applications_sheet = workbook.active
        if applications_sheet is None:
            # Workbook() always creates one active sheet; this only guards mypy's Optional.
            raise ExportError(f"could not write XLSX to {resolved}: workbook has no active sheet")
        applications_sheet.title = "Applications"
        _write_sheet(applications_sheet, df)
        if events is not None:
            events_sheet = workbook.create_sheet("Events")
            _write_sheet(events_sheet, events)
        workbook.save(resolved)
    except (OSError, ValueError, TypeError) as exc:
        raise ExportError(f"could not write XLSX to {resolved}: {exc}") from exc
    return resolved


def _write_sheet(sheet: Worksheet, df: pd.DataFrame) -> None:
    """Write `df` into `sheet`: header, rows, frozen header, autosized columns, autofilter."""
    headers = list(df.columns)
    sheet.append(headers)
    for record in df.itertuples(index=False, name=None):
        sheet.append([_cell_value(value) for value in record])

    sheet.freeze_panes = "A2"

    last_row = sheet.max_row
    last_col_letter = get_column_letter(len(headers)) if headers else "A"
    sheet.auto_filter.ref = f"A1:{last_col_letter}{max(last_row, 1)}"

    for col_index, header in enumerate(headers, start=1):
        widths = [len(str(header))]
        for row in sheet.iter_rows(
            min_row=2, max_row=last_row, min_col=col_index, max_col=col_index
        ):
            cell_text = "" if row[0].value is None else str(row[0].value)
            widths.append(len(cell_text))
        sheet.column_dimensions[get_column_letter(col_index)].width = min(
            max(widths) + 2, _MAX_COLUMN_WIDTH
        )


def _cell_value(value: Any) -> Any:
    """Coerce one pandas cell value into something openpyxl can write.

    Pandas NA/NaT/NaN become ``None``; tz-aware timestamps become ISO-8601 strings
    (openpyxl has no native tz-aware datetime support); everything else passes through.
    """
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value
