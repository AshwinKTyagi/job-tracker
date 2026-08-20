"""Unit tests for jobtrack.export.tabular."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import pytest

from jobtrack.constants import EVENT_COLUMNS, EXPORT_COLUMNS
from jobtrack.errors import ExportError
from jobtrack.export import build_dataframe, build_events_dataframe, write_csv, write_xlsx
from jobtrack.models import ApplicationRow, ApplicationStatus, EventRow, EventType

_APPLIED_AT = datetime(2026, 6, 1, 9, 0, 0, tzinfo=UTC)
_LAST_EVENT_AT = datetime(2026, 6, 10, 15, 30, 0, tzinfo=UTC)
_OCCURRED_AT = datetime(2026, 6, 5, 12, 0, 0, tzinfo=UTC)


def make_application(**overrides: Any) -> ApplicationRow:
    """Build an ApplicationRow with sensible defaults, overriding only what a test needs."""
    defaults: dict[str, Any] = {
        "application_id": "app-0001",
        "company": "Acme Robotics",
        "company_key": "acme robotics",
        "role": "Software Engineer",
        "location": "Remote",
        "ats": "greenhouse",
        "status": ApplicationStatus.INTERVIEWING,
        "applied_at": _APPLIED_AT,
        "last_event_at": _LAST_EVENT_AT,
        "last_event_type": EventType.INTERVIEW,
        "event_count": 3,
        "days_to_first_response": 4,
        "days_since_last_event": 2,
        "needs_review": False,
        "source_thread_ids": ["thread-0001"],
    }
    defaults.update(overrides)
    return ApplicationRow.model_validate(defaults)


def make_event(**overrides: Any) -> EventRow:
    """Build an EventRow with sensible defaults, overriding only what a test needs."""
    defaults: dict[str, Any] = {
        "event_id": 1,
        "application_id": "app-0001",
        "message_id": "msg-0001",
        "event_type": EventType.INTERVIEW,
        "occurred_at": _OCCURRED_AT,
        "confidence": 0.9,
        "needs_review": False,
        "is_overridden": False,
        "subject": "Interview scheduled",
        "from_email": "no-reply@greenhouse.io",
    }
    defaults.update(overrides)
    return EventRow.model_validate(defaults)


# --- build_dataframe --------------------------------------------------------


def test_build_dataframe_column_order_matches_export_columns() -> None:
    df = build_dataframe([make_application()])
    assert list(df.columns) == list(EXPORT_COLUMNS)


def test_build_dataframe_values_round_trip() -> None:
    app = make_application()
    df = build_dataframe([app])
    row = df.iloc[0]
    assert row["application_id"] == app.application_id
    assert row["company"] == app.company
    assert row["role"] == app.role
    assert row["status"] == "interviewing"
    assert row["last_event_type"] == "interview"
    assert row["event_count"] == 3
    assert row["days_to_first_response"] == 4
    assert row["needs_review"] is False or bool(row["needs_review"]) is False
    assert row["applied_at"] == pd.Timestamp(_APPLIED_AT)


def test_build_dataframe_datetimes_are_utc() -> None:
    df = build_dataframe([make_application()])
    assert str(df["applied_at"].dtype) == "datetime64[ns, UTC]"
    assert str(df["last_event_at"].dtype) == "datetime64[ns, UTC]"


def test_build_dataframe_none_fields_become_missing() -> None:
    app = make_application(role=None, location=None, ats=None, days_to_first_response=None)
    df = build_dataframe([app])
    row = df.iloc[0]
    assert pd.isna(row["role"])
    assert pd.isna(row["location"])
    assert pd.isna(row["ats"])
    assert pd.isna(row["days_to_first_response"])


def test_build_dataframe_empty_input_has_correct_columns_and_dtypes() -> None:
    empty_df = build_dataframe([])
    populated_df = build_dataframe([make_application()])

    assert len(empty_df) == 0
    assert list(empty_df.columns) == list(EXPORT_COLUMNS)
    for column in EXPORT_COLUMNS:
        assert empty_df[column].dtype == populated_df[column].dtype, column


def test_build_dataframe_preserves_input_order() -> None:
    app_a = make_application(application_id="app-a")
    app_b = make_application(application_id="app-b")
    df = build_dataframe([app_b, app_a])
    assert list(df["application_id"]) == ["app-b", "app-a"]


# --- build_events_dataframe -------------------------------------------------


def test_build_events_dataframe_columns() -> None:
    df = build_events_dataframe([make_event()])
    assert list(df.columns) == list(EVENT_COLUMNS)
    assert list(df.columns) == [
        "application_id",
        "message_id",
        "event_type",
        "occurred_at",
        "confidence",
        "needs_review",
        "subject",
    ]


def test_build_events_dataframe_values() -> None:
    event = make_event()
    df = build_events_dataframe([event])
    row = df.iloc[0]
    assert row["application_id"] == event.application_id
    assert row["message_id"] == event.message_id
    assert row["event_type"] == "interview"
    assert row["confidence"] == pytest.approx(0.9)
    assert row["subject"] == event.subject


def test_build_events_dataframe_null_application_id() -> None:
    event = make_event(application_id=None, event_type=EventType.UNKNOWN)
    df = build_events_dataframe([event])
    assert pd.isna(df.iloc[0]["application_id"])


def test_build_events_dataframe_empty_input_has_correct_columns_and_dtypes() -> None:
    empty_df = build_events_dataframe([])
    populated_df = build_events_dataframe([make_event()])

    assert len(empty_df) == 0
    assert list(empty_df.columns) == list(EVENT_COLUMNS)
    for column in EVENT_COLUMNS:
        assert empty_df[column].dtype == populated_df[column].dtype, column


# --- determinism -------------------------------------------------------------


def test_build_dataframe_is_deterministic() -> None:
    apps = [make_application(), make_application(application_id="app-0002")]
    first = build_dataframe(apps)
    second = build_dataframe(apps)
    pd.testing.assert_frame_equal(first, second)


# --- write_csv ---------------------------------------------------------------


def test_write_csv_round_trip(tmp_path: Path) -> None:
    df = build_dataframe([make_application(), make_application(application_id="app-0002")])
    out_path = tmp_path / "applications.csv"

    result = write_csv(df, out_path)

    assert result == out_path.resolve()
    assert result.is_file()
    read_back = pd.read_csv(result)
    assert list(read_back.columns) == list(EXPORT_COLUMNS)
    assert len(read_back) == 2
    assert set(read_back["application_id"]) == {"app-0001", "app-0002"}
    # ISO-8601, no index column leaked in.
    assert "Unnamed: 0" not in read_back.columns
    assert str(read_back.loc[0, "applied_at"]).startswith("2026-06-01T09:00:00")


def test_write_csv_utf8_and_no_index(tmp_path: Path) -> None:
    df = build_dataframe([make_application(company="Acme Robotics — Zürich")])
    out_path = tmp_path / "unicode.csv"

    write_csv(df, out_path)

    text = out_path.read_text(encoding="utf-8")
    assert "Acme Robotics — Zürich" in text
    assert text.startswith("application_id,")  # no leading index column


def test_write_csv_empty_dataframe(tmp_path: Path) -> None:
    df = build_dataframe([])
    out_path = tmp_path / "empty.csv"

    result = write_csv(df, out_path)

    read_back = pd.read_csv(result)
    assert list(read_back.columns) == list(EXPORT_COLUMNS)
    assert len(read_back) == 0


def test_write_csv_unwritable_directory_raises_export_error(tmp_path: Path) -> None:
    df = build_dataframe([make_application()])
    bad_path = tmp_path / "does-not-exist" / "applications.csv"

    with pytest.raises(ExportError):
        write_csv(df, bad_path)


def test_write_csv_path_is_a_directory_raises_export_error(tmp_path: Path) -> None:
    df = build_dataframe([make_application()])
    directory = tmp_path / "a_directory"
    directory.mkdir()

    with pytest.raises(ExportError):
        write_csv(df, directory)


# --- write_xlsx ----------------------------------------------------------------


def test_write_xlsx_applications_only_round_trip(tmp_path: Path) -> None:
    df = build_dataframe([make_application(), make_application(application_id="app-0002")])
    out_path = tmp_path / "applications.xlsx"

    result = write_xlsx(df, out_path)

    assert result == out_path.resolve()
    workbook = openpyxl.load_workbook(result)
    assert workbook.sheetnames == ["Applications"]
    sheet = workbook["Applications"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == list(EXPORT_COLUMNS)
    first_data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert first_data_row[0] == "app-0001"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


def test_write_xlsx_with_events_adds_second_sheet(tmp_path: Path) -> None:
    apps_df = build_dataframe([make_application()])
    events_df = build_events_dataframe([make_event()])
    out_path = tmp_path / "with_events.xlsx"

    write_xlsx(apps_df, out_path, events=events_df)

    workbook = openpyxl.load_workbook(out_path)
    assert workbook.sheetnames == ["Applications", "Events"]
    events_sheet = workbook["Events"]
    header = [cell.value for cell in next(events_sheet.iter_rows(min_row=1, max_row=1))]
    assert header == list(EVENT_COLUMNS)
    assert events_sheet.freeze_panes == "A2"


def test_write_xlsx_autosizes_columns(tmp_path: Path) -> None:
    long_company = "A" * 80
    df = build_dataframe([make_application(company=long_company)])
    out_path = tmp_path / "wide.xlsx"

    write_xlsx(df, out_path)

    workbook = openpyxl.load_workbook(out_path)
    sheet = workbook["Applications"]
    company_col_letter = "B"  # company is EXPORT_COLUMNS[1]
    width = sheet.column_dimensions[company_col_letter].width
    assert width is not None
    assert width > len("company")
    assert width <= 62  # capped, never unbounded


def test_write_xlsx_empty_dataframe_still_writes_header(tmp_path: Path) -> None:
    df = build_dataframe([])
    out_path = tmp_path / "empty.xlsx"

    write_xlsx(df, out_path)

    workbook = openpyxl.load_workbook(out_path)
    sheet = workbook["Applications"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == list(EXPORT_COLUMNS)
    assert sheet.max_row == 1


def test_write_xlsx_none_values_are_blank_cells(tmp_path: Path) -> None:
    app = make_application(role=None, location=None, days_to_first_response=None)
    df = build_dataframe([app])
    out_path = tmp_path / "nulls.xlsx"

    write_xlsx(df, out_path)

    workbook = openpyxl.load_workbook(out_path)
    sheet = workbook["Applications"]
    role_idx = EXPORT_COLUMNS.index("role")
    row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert row[role_idx] is None


def test_write_xlsx_unwritable_directory_raises_export_error(tmp_path: Path) -> None:
    df = build_dataframe([make_application()])
    bad_path = tmp_path / "does-not-exist" / "applications.xlsx"

    with pytest.raises(ExportError):
        write_xlsx(df, bad_path)


def test_write_xlsx_path_is_a_directory_raises_export_error(tmp_path: Path) -> None:
    df = build_dataframe([make_application()])
    directory = tmp_path / "a_directory"
    directory.mkdir()

    with pytest.raises(ExportError):
        write_xlsx(df, directory)
